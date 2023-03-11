from bs4 import BeautifulSoup
import openpyxl
import requests

def get_homes(soup):
    homes = []
    home_containers = soup.find_all("div", class_="p-card-facility")
    for count, home_container in enumerate(home_containers):
        if count >= 20: break
        name = home_container.find("h3").text.replace("\n", "")
        url = home_container.find("a").get("href")
        type = home_container.find("p", class_="p-card-facility__type").text
        visit_acception = home_container.find("span", class_="p-card-facility__visit-acception").text
        rating = get_rating(home_container)
        dist_from_station = home_container.find("ul", class_="inline-list shisetsu-line-list").text.replace("\n", ", ")
        description = home_container.find("p", class_="p-card-facility__sales_copy").text
        table_info = get_table_info(home_container)
        homes.append([name, url, type, visit_acception, rating, dist_from_station, description] + table_info)
    return homes

def get_soup(url):
    try: 
        soup = BeautifulSoup(requests.get(url).text, "html.parser")
        return soup
    except:
        print("URL Error: " + url)
        return None
    
def get_table_info(home_container):
    table = home_container.find("table")
    tds = table.find_all("td")
    output = []
    for i, td in enumerate(tds):
        text = arrange_text_format(i, td.text)
        output.append(text)
    return output

def get_rating(home_container):
    rating = home_container.find("div", class_="p-card-facility__head-btm").text.strip()[:4]
    if "部屋" in rating:
        return "記載なし"
    return rating

def arrange_text_format(i, text):
    text = text.replace("\n", " ").replace("\t", "").strip()
    if i==4 and "要介護者" not in text:
        return "記載なし"
    if i==5:
        text = text.replace("施設に直接お問い合わせ", "").replace("みんなの介護を見たと必ずお伝えください。", "").strip()
        if text.find("0") == -1:
            return "記載なし"
    return text

def write_to_excel(courses_data, sheet_name):
    wb = openpyxl.load_workbook('PRTIMES.xlsx')
    sheet = wb.create_sheet(sheet_name)
    for i, key in enumerate(keys):
        sheet.cell(row=1, column=i+1, value=key)
    for i in range(len(courses_data)):
        for j in range(len(courses_data[i])):
            sheet.cell(row=i+2, column=j+1, value=courses_data[i][j])
    wb.save('PRTIMES.xlsx')
    print("complete")

# Run
url = "https://www.minnanokaigo.com/search/premium/"
resoponse = requests.get(url)
keys = ["name", "url", "type", "visit_acception", "rating", "dist_from_station", "description", "price", "address", "established_date", "capacity", "ppl_per_staff", "tel", "company"]
homes = []

for page_num in range(1, 42):
    soup = get_soup(url + str(page_num) + "/")
    homes += get_homes(soup)
    print("page " + str(page_num) + " done")

write_to_excel(homes, "高級な老人ホーム")

