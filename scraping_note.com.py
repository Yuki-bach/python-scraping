from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
import time
import openpyxl

"""
Scraping note.com
 1. Click "もっと見る" button
 2. Get article's data (title, link, date)
"""

driver = webdriver.Chrome()
url = 'https://note.com/buysell/all'
wait = WebDriverWait(driver, 10)
driver.get(url)

# Click "もっと見る" button
while True:
    try:
        time.sleep(3)
        readMoreBt = driver.find_element(By.CSS_SELECTOR, '#__layout > div > div:nth-child(1) > main > div > div.t-newCreator__wrapper > div.t-newCreator__content > div > section > div.o-contentNote__forPc > div > div > div > button') #もっと見るボタン
        readMoreBt.click()
        print("clicked")
    except (NoSuchElementException) as e:
        print('Clicked All "Read More" Button')
        break

# Get title , link, and date
time.sleep(10)
html = driver.page_source.encode('utf-8')
soup = BeautifulSoup(html, "lxml")

divs = soup.find_all("div",class_="renewal-c-card")
export_list = []

for div in divs:
    title = div.find(class_="renewal-p-cardItem__title").text
    link = div.find_all(class_="a-link")[2].get("href")
    link = "note.com" + link
    date = div.find("time").get("datetime")[:10]
    temp = []
    temp.append(title)
    temp.append(link)
    temp.append(date)
    export_list.append(temp)

driver.close()
driver.quit()

# Write to excel
wb = openpyxl.load_workbook('PRTIMES.xlsx')
sheet = wb['Sheet1']
for i in range(len(export_list)):
  sheet.cell(row=i+1, column=1, value=export_list[i][0])
  sheet.cell(row=i+1, column=2, value=export_list[i][1])
  sheet.cell(row=i+1, column=3, value=export_list[i][2])

wb.save('PRTIMES.xlsx')
print("complete")