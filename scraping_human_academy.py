from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl
import re

def get_courses_by_category(soup, category):
    courses_data = []
    urls_titles = get_courses_url_title(soup, category)
    for [url, title] in urls_titles:
        course_data = get_course_data(url, title)
        if (course_data is not None):
            courses_data.append(course_data)
    print(courses_data)
    write_to_excel(courses_data, category)

def get_courses_url_title(soup, course_category):
    hobby = soup.find("div", id=course_category)
    hobby_a = hobby.find_all("a")
    course_list = []
    for a in hobby_a:
        title = a.text
        url = a.get("href")
        if "http" not in url:
            url = "https://www.tanomana.com" + a.get("href")
        course_list.append([url, title])
    return course_list

def get_course_data(url, title):
    soup = get_soup(url)
    
    course_wrapper = get_course_wrapper(soup)
    can_get_course_wrapper = course_wrapper != None
    if can_get_course_wrapper:
        price = get_price(course_wrapper)
        teacher = get_teacher(soup)
        period = get_period(course_wrapper)
        return [title, url, price, teacher, period]
    return [title, url, "", "", ""]

def get_soup(url):
    try: 
        driver.get(url)
        html = driver.page_source.encode('utf-8')
        soup = BeautifulSoup(html, "lxml")

        if soup.find("p", class_="error404__txt") is None:
            return soup
    except:
        print("URL Error: " + url)
        return None

def get_course_wrapper(soup):
    class_names = ["coursebox", "pack train vid", "layoutp3", "case bg", "container bg-greenLight p-5", "container bg-yellow05 p-5", "container bg-orangeLight02 p-5", "container bg-lightBlue p-5", "top__contentInner"] 
    for class_name in class_names:
        return get_element_helper(soup, "div", class_name)
    return None

def get_period(course_wrapper):
    class_names = ["period", "mb10", "item fl_l", "price-exam-list clm6 clearfix mb-5", "price-exam-list clearfix mb-5", "clearfix mx-auto courseMat"]
    period_text = get_element_text(course_wrapper, "ul", class_names)
    if period_text is not None:
        return period_text.replace('\n', ', ')
    return ""

def get_price(course_wrapper):
    class_names_set = [["div", ["sum", "price", "item fl_l sum", "price no2"]],
                        ["span", ["f300"]],
                          ["p", ["top_price"]]]
    for tag, class_names in class_names_set:
        price_text = get_element_text(course_wrapper, tag, class_names)
        if price_text != "":
            return re.search(r'\d+,\d+', price_text).group()
    return ""

def get_teacher(soup):
    container = get_element_helper(soup, "div", "box_last")
    div = get_element_helper(container, "div", "img_area")
    if div is not None:
        return div.find("span").text
    return ""

def get_element_text(wrapper, tag, class_names):
    if wrapper is None: return ""
    for class_name in class_names:
        if (wrapper.find(tag, class_=class_name) is not None):
            return wrapper.find(tag, class_=class_name).text
    return ""

def get_element_helper(wrapper, tag, class_name):
    if wrapper is None: return None
    if (wrapper.find(tag, class_=class_name) is not None):
        return wrapper.find(tag, class_=class_name)
    return None

def write_to_excel(courses_data, sheet_name):
    wb = openpyxl.load_workbook('PRTIMES.xlsx')
    sheet = wb.create_sheet(sheet_name)
    for i in range(len(courses_data)):
        for j in range(5):
            sheet.cell(row=i+2, column=j+1, value=courses_data[i][j])
    wb.save('PRTIMES.xlsx')
    print("complete")


# Set driver
driver = webdriver.Chrome()
wait = WebDriverWait(driver, 5)
driver.get('https://www.tanomana.com/hpgen/HPB/entries/20.html')
html = driver.page_source.encode('utf-8')
soup = BeautifulSoup(html, "lxml")

# Get courses
# categories = ["hobby", "language", "food", "zaitaku", "web", "business", "care", "beauty", "mental", "medical", "deco", "nail", "animal"]
# for category in categories:
#     get_courses_by_category(soup, category)

driver.close()
driver.quit()